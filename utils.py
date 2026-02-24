from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

def parse_summary_to_dict(summary_text):
    """
    Parses the structured summary text into a dictionary of sections.
    """
    sections = {}
    current_section = None
    buffer = []

    # Clean up the summary text first
    lines = summary_text.split('\n')

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Check if line is a header (starts and ends with **)
        # e.g. **PATIENT INFORMATION:**
        if line.startswith('**') and (line.endswith('**') or line.endswith(':')):
            # Save the previous section if it exists
            if current_section:
                sections[current_section] = "\n".join(buffer).strip()
            
            # Start new section
            # remove **, :, and spaces
            current_section = line.replace('*', '').replace(':', '').strip()
            buffer = []
        else:
            # Add line to current section buffer
            buffer.append(line)

    # Save the last section
    if current_section and buffer:
        sections[current_section] = "\n".join(buffer).strip()
    elif not sections and buffer:
        # Fallback if no headers found
        sections["Summary"] = "\n".join(buffer).strip()

    return sections

def export_to_excel(filename, transcript, entities, summary):
    wb = Workbook()
    ws = wb.active
    ws.title = "Clinical Report"

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 80

    # Header Style
    header_font = Font(bold=True, size=12)
    
    # 1. Transcript & Raw Data
    ws.append(["Field", "Value"])
    for cell in ws[1]:
        cell.font = header_font

    ws.append(["Transcript", transcript])
    # Format symptoms with negation
    symptom_list = []
    for s in entities.get("symptoms", []):
        if isinstance(s, dict):
            text = s.get("text", "")
            if s.get("negated"):
                text += " (Negated)"
            symptom_list.append(text)
        else:
            symptom_list.append(str(s))
            
    # Format medications with negation
    med_list = []
    for m in entities.get("medications", []):
        if isinstance(m, dict):
            text = m.get("text", "")
            if m.get("negated"):
                text += " (Negated)"
            med_list.append(text)
        else:
            med_list.append(str(m))

    ws.append(["Extracted Symptoms", ", ".join(symptom_list)])
    ws.append(["Extracted Medications", ", ".join(med_list)])
    ws.append(["Family & Social Details", "\n".join(entities.get("family_friends", []))])
    ws.append([]) # spacer

    # 2. Clinical Summary Breakdown
    summary_sections = parse_summary_to_dict(summary)
    
    # Add a section header row
    ws.append(["--- CLINICAL SUMMARY ---", "--- COMPREHENSIVE BREAKDOWN ---"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, italic=True)

    for section, content in summary_sections.items():
        ws.append([section, content])
        
        # Style the row
        # Column A (Header): Bold, Top/Left aligned
        cell_key = ws.cell(row=ws.max_row, column=1)
        cell_key.font = Font(bold=True)
        cell_key.alignment = Alignment(vertical='top', wrap_text=True)

        # Column B (Content): Wrap text, Top/Left aligned
        cell_val = ws.cell(row=ws.max_row, column=2)
        cell_val.alignment = Alignment(vertical='top', wrap_text=True)

    wb.save(filename)

